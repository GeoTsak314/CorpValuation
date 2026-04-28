# CORP Valuation app v5.7 by George Tsakalos


import configparser
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import webbrowser
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

try:
    import openpyxl
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, BaseDocTemplate, PageTemplate, Frame, NextPageTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics import renderPDF
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False




# CFG Defaults 
APP_TITLE = "CORP Valuation app v5.7 (by G.Tsakalos)"
DEFAULT_DB = "corp_values.sqlite"
DEFAULT_CFG = "app.cfg"
DEFAULT_INDEX_MAP = "IndexMap.jpg"

COMPANY_FIELDS = [
    ("name", "Επωνυμία"),
    ("afm", "ΑΦΜ"),
    ("gemi", "ΓΕΜΗ Αρ."),
    ("gemi_url", "ΓΕΜΗ URL"),
    ("company_url", "URL εταιρείας"),
    ("notes", "Σημειώσεις"),
]

BALANCE_NUMERIC_FIELDS = [
    "fixed_tangible", "intangible_assets", "subsidiary_investments", "other_company_investments",
    "right_of_use", "investment_property", "other_long_term_receivables", "non_current_diff",
    "inventory", "trade_receivables", "other_short_term_receivables", "other_current_assets",
    "cash_and_equivalents", "current_diff", "share_capital", "share_premium", "reserves",
    "retained_earnings", "equity_diff", "long_term_liabilities", "short_term_liabilities", "liabilities_diff"
]

INCOME_NUMERIC_FIELDS = [
    "net_sales", "state_grants", "cost_of_sales",
    "admin_expenses", "selling_expenses", "research_expenses", "provisions_expenses",
    "other_operating_expenses", "inventory_revaluation_losses", "other_expenses",
    "other_operating_income", "inventory_revaluation_gains", "other_income",
    "depr_tangible", "depr_intangible",
    "dividend_income", "financial_expenses", "other_financial_losses",
    "financial_income", "other_financial_gains",
    "income_taxes", "ebitda_manual"
]

BALANCE_IMPORT_MAP = {
    5: "fixed_tangible",
    6: "intangible_assets",
    7: "subsidiary_investments",
    8: "other_company_investments",
    9: "right_of_use",
    10: "investment_property",
    11: "other_long_term_receivables",
    12: "non_current_diff",
    15: "inventory",
    16: "trade_receivables",
    17: "other_short_term_receivables",
    18: "other_current_assets",
    19: "cash_and_equivalents",
    20: "current_diff",
    28: "share_capital",
    29: "share_premium",
    30: "reserves",
    31: "retained_earnings",
    32: "equity_diff",
    35: "long_term_liabilities",
    36: "short_term_liabilities",
    37: "liabilities_diff",
}

INCOME_IMPORT_MAP = {
    3: "net_sales",
    4: "state_grants",
    5: "cost_of_sales",
    9: "other_operating_income",
    10: "inventory_revaluation_gains",
    11: "other_income",
    13: "admin_expenses",
    14: "selling_expenses",
    15: "research_expenses",
    16: "provisions_expenses",
    17: "other_operating_expenses",
    18: "inventory_revaluation_losses",
    19: "other_expenses",
    21: "ebitda_manual",
    23: "depr_tangible",
    24: "depr_intangible",
    27: "financial_income",
    28: "other_financial_gains",
    29: "dividend_income",
    30: "other_financial_losses",
    31: "financial_expenses",
    34: "income_taxes",
}

BALANCE_LAYOUT = [
    ("header", "ΕΝΕΡΓΗΤΙΚΟ", None),
    ("field", "Ενσώματα πάγια", "fixed_tangible"),
    ("field", "Άυλα περιουσιακά στοιχεία", "intangible_assets"),
    ("field", "Επενδύσεις σε θυγατρικές επιχειρήσεις", "subsidiary_investments"),
    ("field", "Επενδύσεις σε λοιπές εταιρείες", "other_company_investments"),
    ("field", "Δικαιωματα χρήσης", "right_of_use"),
    ("field", "Επενδύσεις σε ακίνητα", "investment_property"),
    ("field", "Λοιπές μακροπρόθεσμες απαιτήσεις", "other_long_term_receivables"),
    ("diff", "Διάφορα Μη-κυκλοφορούντα", "non_current_diff"),
    ("auto", "Μη-κυκλοφορούν Ενεργητικό", "non_current_assets"),
    ("space", "", None),
    ("field", "Αποθέματα", "inventory"),
    ("field", "Απαιτήσεις από πελάτες και λοιπές εμπορικές απαιτήσεις", "trade_receivables"),
    ("field", "Λοιπές βραχυπρόθεσμες απαιτήσεις", "other_short_term_receivables"),
    ("field", "Λοιποί λογαριασμοί ενεργητικού", "other_current_assets"),
    ("field", "Ταμειακά διαθέσιμα και ισοδύναμα", "cash_and_equivalents"),
    ("diff", "Διάφορα Κυκλοφορούντα", "current_diff"),
    ("auto", "Κυκλοφορούν Ενεργητικό", "current_assets"),
    ("space", "", None),
    ("auto_header", "Σύνολο ΕΝΕΡΓΗΤΙΚΟΥ", "total_assets"),
    ("space", "", None),
    ("header", "ΠΑΘΗΤΙΚΟ", None),
    ("field", "Μ/Κ", "share_capital"),
    ("field", "Διαφορά υπέρ το άρτιο", "share_premium"),
    ("field", "Αποθεματικά κεφάλαια", "reserves"),
    ("field", "Αποτελέσματα εις νέο", "retained_earnings"),
    ("diff", "Διάφορα Ι/Κ", "equity_diff"),
    ("auto", "Ι/Κ (Ίδια Κεφάλαια)", "equity"),
    ("space", "", None),
    ("field", "Μακροπρόθεσμες υποχρεώσεις", "long_term_liabilities"),
    ("field", "Βραχυπρόθεσμες υποχρεώσεις", "short_term_liabilities"),
    ("diff", "Διάφορες Υποχρεώσεις", "liabilities_diff"),
    ("auto", "Υποχρεώσεις", "liabilities"),
    ("space", "", None),
    ("auto_header", "Σύνολο ΠΑΘΗΤΙΚΟΥ", "total_liabilities_equity"),
]

INCOME_LAYOUT = [
    ("field", "Κύκλος εργασιών (net sales)", "net_sales"),
    ("field", "Επιχορηγήσεις Δημοσίου", "state_grants"),
    ("field", "Κόστος πωλήσεων", "cost_of_sales"),
    ("auto", "Μικτό κέρδος (/ζημιά)", "gross_profit"),
    ("space", "", None),

    ("header", "ΕΞΟΔΑ", None),
    ("field", "Έξοδα διοίκησης", "admin_expenses"),
    ("field", "Έξοδα διάθεσης", "selling_expenses"),
    ("field", "Έξοδα έρευνας", "research_expenses"),
    ("field", "Έξοδα προβλέψεων", "provisions_expenses"),
    ("field", "Λοιπά έξοδα εκμετάλλευσης", "other_operating_expenses"),
    ("field", "Ζημιές από επιμέτρηση αποθεμάτων", "inventory_revaluation_losses"),
    ("field", "Άλλα έξοδα", "other_expenses"),
    ("space", "", None),

    ("header", "ΕΣΟΔΑ", None),
    ("field", "Λοιπά έσοδα εκμετάλλευσης", "other_operating_income"),
    ("field", "Κέρδη από επιμέτρηση αποθεμάτων", "inventory_revaluation_gains"),
    ("field", "Άλλα έσοδα", "other_income"),
    ("auto", "Αποτελέσματα εκμετάλλ. προ φόρων, χρημ/κών και επενδυτ. Αποτελεσμάτων (EBIT)", "ebit"),
    ("field_bold", "Αποτελέσματα εκμετάλλ. προ φόρων, χρημ/κών και αποσβέσεων (EBITDA)", "ebitda_manual"),
    ("space", "", None),

    ("field", "Αποσβέσεις χρήσης - Ενσώματα πάγια", "depr_tangible"),
    ("field", "Αποσβέσεις χρήσης - Άυλα πάγια", "depr_intangible"),
    ("auto", "Depreciations (Αποσβέσεις Χρήσης)", "depreciation"),
    ("space", "", None),

    ("field", "Χρηματοοικονομικά έξοδα", "financial_expenses"),
    ("field", "Λοιπές Ζημιές", "other_financial_losses"),
    ("space", "", None),
    ("field", "Έσοδα από μερίσματα", "dividend_income"),
    ("field", "Χρηματοοικονομικά έσοδα", "financial_income"),
    ("field", "Λοιπά Κέρδη", "other_financial_gains"),
    ("auto", "Αποτελέσματα προ φόρων", "pbt"),
    ("space", "", None),

    ("field", "Φόροι Εισοδήματος", "income_taxes"),
    ("auto", "Αποτελέσματα χρήσης μετά φόρων", "pat"),
]

RATIO_ORDER = [
    "Profit Margin (Περιθώριο Καθαρού Κέρδους)",
    "Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)",
    "ROA (Απόδοση Συνόλου Ενεργητικού)",
    "Financial Leverage (Χρημ/ική Μόχλευση)",
    "ROE (Απόδοση Ι/Κ)",
    "Current Ratio (Κεφ./Κίνησης)",
    "Quick Ratio (Ρευστότητας)",
    "Inventory Turnover Ratio (Κυκλοφορ. Ταχύτ. Αποθεμάτων σε φορές)",
    "Inventory Days (Παραμονή Αποθεμάτων σε ημέρες)",
    "Receivable Turnover Ratio (Κυκλοφορ. Ταχύτ. Απαιτήσεων σε φορές)",
    "Receivable Days (Παραμονή Απαιτήσεων σε ημέρες)",
    "Payable Turnover Ratio (Κυκλοφορ. Ταχύτ. Υποχρεώσεων σε φορές)",
    "Payable Days (Παραμονή Βραχυπρ. Υποχρεώσεων σε ημέρες)",
    "Operating Cycle (Λειτουργικός Κύκλος σε ημέρες)",
    "Working Capital Requirements (Ανάγκες σε Κ/Κ)",
    "Fixed Asset Turnover (Κυκλοφορ. Ταχύτ. Πάγιου Ενεργητικού σε φορές)",
    "Total Asset Turnover (Κυκλοφορ. Ταχύτ. Ενεργητικού σε φορές)",
    "EBIT",
    "EBITDA",
    "Depreciations (Αποσβέσεις Χρήσης)",
]

RATIO_GROUPS = [
    ("Επενδυτικοί", [
        "Profit Margin (Περιθώριο Καθαρού Κέρδους)",
        "Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)",
        "ROA (Απόδοση Συνόλου Ενεργητικού)",
        "Financial Leverage (Χρημ/ική Μόχλευση)",
        "ROE (Απόδοση Ι/Κ)",
    ]),
    ("Ρευστότητας", [
        "Current Ratio (Κεφ./Κίνησης)",
        "Quick Ratio (Ρευστότητας)",
    ]),
    ("Επιχ. Αποδοτικότητας", [
        "Inventory Turnover Ratio (Κυκλοφορ. Ταχύτ. Αποθεμάτων σε φορές)",
        "Inventory Days (Παραμονή Αποθεμάτων σε ημέρες)",
        "Receivable Turnover Ratio (Κυκλοφορ. Ταχύτ. Απαιτήσεων σε φορές)",
        "Receivable Days (Παραμονή Απαιτήσεων σε ημέρες)",
        "Payable Turnover Ratio (Κυκλοφορ. Ταχύτ. Υποχρεώσεων σε φορές)",
        "Payable Days (Παραμονή Βραχυπρ. Υποχρεώσεων σε ημέρες)",
        "Operating Cycle (Λειτουργικός Κύκλος σε ημέρες)",
        "Working Capital Requirements (Ανάγκες σε Κ/Κ)",
        "Fixed Asset Turnover (Κυκλοφορ. Ταχύτ. Πάγιου Ενεργητικού σε φορές)",
        "Total Asset Turnover (Κυκλοφορ. Ταχύτ. Ενεργητικού σε φορές)",
        "EBIT",
        "EBITDA",
        "Depreciations (Αποσβέσεις Χρήσης)",
    ]),
]


def iter_grouped_ratios():
    for group_name, ratio_names in RATIO_GROUPS:
        yield ("__GROUP__", group_name)
        for ratio_name in ratio_names:
            yield ("ratio", ratio_name)

PERCENT_RATIOS = {
    "Profit Margin (Περιθώριο Καθαρού Κέρδους)",
    "ROA (Απόδοση Συνόλου Ενεργητικού)",
    "ROE (Απόδοση Ι/Κ)",}

PLAIN_RATIO_RATIOS = {
    "Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)",
    "Financial Leverage (Χρημ/ική Μόχλευση)",
    "Current Ratio (Κεφ./Κίνησης)",
    "Quick Ratio (Ρευστότητας)",
    "Inventory Turnover Ratio (Κυκλοφορ. Ταχύτ. Αποθεμάτων σε φορές)",
    "Receivable Turnover Ratio (Κυκλοφορ. Ταχύτ. Απαιτήσεων σε φορές)",
    "Payable Turnover Ratio (Κυκλοφορ. Ταχύτ. Υποχρεώσεων σε φορές)",
    "Fixed Asset Turnover (Κυκλοφορ. Ταχύτ. Πάγιου Ενεργητικού σε φορές)",
    "Total Asset Turnover (Κυκλοφορ. Ταχύτ. Ενεργητικού σε φορές)",}

AMOUNT_RATIOS = {
    "Working Capital Requirements (Ανάγκες σε Κ/Κ)",
    "EBIT",
    "EBITDA",
    "Depreciations (Αποσβέσεις Χρήσης)",}


DECIMAL4_RATIOS = {
    "Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)",}

def format_ratio_display(ratio_name: str, value):
    if value is None:
        return "-"
    if ratio_name in PERCENT_RATIOS:
        return fmt_pct(value)
    if ratio_name in DECIMAL4_RATIOS:
        return fmt_num4(value)
    return fmt_num(value)

def is_percent_ratio(ratio_name: str) -> bool:
    return ratio_name in PERCENT_RATIOS







# METHODS
class ImportCanceled(Exception):
    pass


def resource_path(relative_path: str) -> str:
    if getattr(sys, "frozen", False):
        base_path = Path(sys.executable).resolve().parent
    else:
        base_path = Path(__file__).resolve().parent
    return str(base_path / relative_path)


def open_with_default_app(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    if os.name == "nt":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


def normalize_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if not url.lower().startswith(("http://", "https://")):
        url = "https://" + url
    return url


def open_url(url: str):
    url = normalize_url(url)
    if not url:
        raise ValueError("Δεν υπάρχει URL.")
    webbrowser.open(url)



def format_thousands_dot(value):
    if value in (None, ""):
        return ""
    try:
        v = float(value)
        negative = v < 0
        v = abs(v)
        if abs(v - round(v)) < 1e-9:
            s = f"{int(round(v)):,}".replace(",", ".")
        else:
            s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"-{s}" if negative else s
    except Exception:
        return str(value)


def strip_thousands_format(text):
    text = (text or "").strip()
    if not text:
        return ""
    return text.replace(".", "").replace(",", ".")


def format_entry_thousands(entry):
    try:
        value = entry.get().strip()
        if not value:
            return
        parsed = safe_float(value)
        entry.delete(0, "end")
        entry.insert(0, format_thousands_dot(parsed))
    except Exception:
        pass


def unformat_entry_thousands(entry):
    try:
        value = entry.get().strip()
        if not value:
            return
        entry.delete(0, "end")
        entry.insert(0, strip_thousands_format(value))
    except Exception:
        pass

def safe_float(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts) > 2 or (len(parts) > 1 and all(len(p) == 3 for p in parts[1:])):
            s = "".join(parts)
    try:
        return float(s)
    except Exception:
        return 0.0


def fmt_num(value):
    try:
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"


def fmt_num4(value):
    try:
        return f"{float(value):,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"


def fmt_pct(value):
    if value is None:
        return "-"
    try:
        return f"{float(value) * 100:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"


def div(a, b):
    if b in (0, None):
        return None
    try:
        return a / b
    except Exception:
        return None


def ensure_default_cfg(cfg_path: str):
    cfg = configparser.ConfigParser()
    if os.path.exists(cfg_path):
        cfg.read(cfg_path, encoding="utf-8")
    if "app" not in cfg:
        cfg["app"] = {}
    appsec = cfg["app"]
    appsec.setdefault("db_path", resource_path(DEFAULT_DB))
    appsec.setdefault("index_map_path", resource_path(DEFAULT_INDEX_MAP))
    with open(cfg_path, "w", encoding="utf-8") as f:
        cfg.write(f)
    return cfg


def get_cfg_value(cfg_path: str, section: str, key: str, default: str):
    cfg = ensure_default_cfg(cfg_path)
    return cfg.get(section, key, fallback=default)


def get_reportlab_font_paths():
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
    ]
    for regular, bold in candidates:
        if os.path.exists(regular) and os.path.exists(bold):
            return regular, bold
    return None, None


def register_pdf_fonts():
    if not REPORTLAB_OK:
        return ("Helvetica", "Helvetica-Bold")
    regular, bold = get_reportlab_font_paths()
    if regular and bold:
        try:
            if "corp-Regular" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("corp-Regular", regular))
            if "corp-Bold" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("corp-Bold", bold))
            return ("corp-Regular", "corp-Bold")
        except Exception:
            pass
    return ("Helvetica", "Helvetica-Bold")

def style_excel_header(ws, row_idx: int, fill_color: str = "D9E2F3"):
    if not OPENPYXL_OK:
        return
    fill = openpyxl.styles.PatternFill("solid", fgColor=fill_color)
    font = openpyxl.styles.Font(bold=True)
    align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin", color="999999"),
        right=openpyxl.styles.Side(style="thin", color="999999"),
        top=openpyxl.styles.Side(style="thin", color="999999"),
        bottom=openpyxl.styles.Side(style="thin", color="999999"),
    )
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def autosize_worksheet(ws, min_width: int = 10, max_width: int = 40):
    if not OPENPYXL_OK:
        return
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def format_value_for_excel(ratio_name: str, value):
    if value is None:
        return None
    return value


def add_report_charts_sheet(wb, report_rows: list[dict]):
    if not OPENPYXL_OK or not report_rows:
        return
    try:
        from openpyxl.chart import LineChart, Reference
    except Exception:
        return

    ws = wb.create_sheet("Charts")
    ws.append(["Metric"] + [a["year"] for a in report_rows])
    metric_map = {
        "EBIT": [a["ratios"].get("EBIT") for a in report_rows],
        "ROA": [a["ratios"].get("ROA (Απόδοση Συνόλου Ενεργητικού)") for a in report_rows],
        "ROE": [a["ratios"].get("ROE (Απόδοση Ι/Κ)") for a in report_rows],
    }
    metric_names = ["EBIT", "ROA", "ROE"]
    for metric in metric_names:
        ws.append([metric] + metric_map[metric])

    style_excel_header(ws, 1, "E2F0D9")
    chart_positions = {"EBIT": "A6", "ROA": "J6", "ROE": "A24"}
    for idx, metric in enumerate(metric_names, start=2):
        chart = LineChart()
        chart.title = metric
        chart.style = 2
        chart.y_axis.title = metric
        chart.x_axis.title = "Έτος"
        data = Reference(ws, min_col=2, max_col=1 + len(report_rows), min_row=idx, max_row=idx)
        cats = Reference(ws, min_col=2, max_col=1 + len(report_rows), min_row=1, max_row=1)
        chart.add_data(data, titles_from_data=False, from_rows=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 16
        ws.add_chart(chart, chart_positions[metric])
    autosize_worksheet(ws)



class SQLiteDB:
    def __init__(self, path: str):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.init_schema()

    def close(self):
        self.conn.close()

    def _ensure_column(self, table: str, col: str, sql_type: str, default_sql: str):
        cols = {r["name"] for r in self.conn.execute(f"PRAGMA table_info({table})").fetchall()}
        if col not in cols:
            self.conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {sql_type} DEFAULT {default_sql}")

    def init_schema(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                afm TEXT DEFAULT '',
                gemi TEXT DEFAULT '',
                gemi_url TEXT DEFAULT '',
                company_url TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS balance_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                fixed_tangible REAL DEFAULT 0,
                intangible_assets REAL DEFAULT 0,
                subsidiary_investments REAL DEFAULT 0,
                other_company_investments REAL DEFAULT 0,
                right_of_use REAL DEFAULT 0,
                investment_property REAL DEFAULT 0,
                other_long_term_receivables REAL DEFAULT 0,
                non_current_diff REAL DEFAULT 0,
                inventory REAL DEFAULT 0,
                trade_receivables REAL DEFAULT 0,
                other_short_term_receivables REAL DEFAULT 0,
                other_current_assets REAL DEFAULT 0,
                cash_and_equivalents REAL DEFAULT 0,
                current_diff REAL DEFAULT 0,
                share_capital REAL DEFAULT 0,
                share_premium REAL DEFAULT 0,
                reserves REAL DEFAULT 0,
                retained_earnings REAL DEFAULT 0,
                equity_diff REAL DEFAULT 0,
                long_term_liabilities REAL DEFAULT 0,
                short_term_liabilities REAL DEFAULT 0,
                liabilities_diff REAL DEFAULT 0,
                comments TEXT DEFAULT '',
                updated_at TEXT,
                UNIQUE(company_id, year),
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS income_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                net_sales REAL DEFAULT 0,
                state_grants REAL DEFAULT 0,
                cost_of_sales REAL DEFAULT 0,
                admin_expenses REAL DEFAULT 0,
                selling_expenses REAL DEFAULT 0,
                research_expenses REAL DEFAULT 0,
                provisions_expenses REAL DEFAULT 0,
                other_operating_expenses REAL DEFAULT 0,
                inventory_revaluation_losses REAL DEFAULT 0,
                other_expenses REAL DEFAULT 0,
                other_operating_income REAL DEFAULT 0,
                inventory_revaluation_gains REAL DEFAULT 0,
                other_income REAL DEFAULT 0,
                depr_tangible REAL DEFAULT 0,
                depr_intangible REAL DEFAULT 0,
                financial_expenses REAL DEFAULT 0,
                other_financial_losses REAL DEFAULT 0,
                financial_income REAL DEFAULT 0,
                other_financial_gains REAL DEFAULT 0,
                other_financial_results REAL DEFAULT 0,
                dividend_income REAL DEFAULT 0,
                income_taxes REAL DEFAULT 0,
                ebitda_manual REAL DEFAULT 0,
                comments TEXT DEFAULT '',
                updated_at TEXT,
                UNIQUE(company_id, year),
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ratio_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                ratio_name TEXT NOT NULL,
                note TEXT DEFAULT '',
                updated_at TEXT,
                UNIQUE(company_id, ratio_name),
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            )
            """
        )
        self.conn.commit()

        for col in ["gemi_url", "company_url"]:
            self._ensure_column("companies", col, "TEXT", "''")
        for col in ["non_current_diff", "current_diff", "equity_diff", "liabilities_diff"]:
            self._ensure_column("balance_sheets", col, "REAL", "0")
        self._ensure_column("balance_sheets", "comments", "TEXT", "''")
        for col in [
            "research_expenses", "provisions_expenses", "inventory_revaluation_losses",
            "inventory_revaluation_gains", "other_financial_losses", "other_financial_gains",
            "other_expenses", "ebitda_manual"
        ]:
            self._ensure_column("income_sheets", col, "REAL", "0")
        self._ensure_column("income_sheets", "comments", "TEXT", "''")
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ratio_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                ratio_name TEXT NOT NULL,
                note TEXT DEFAULT '',
                updated_at TEXT,
                UNIQUE(company_id, ratio_name),
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            )
            """
        )
        self.conn.commit()

    def row_to_dict(self, row):
        return dict(row) if row is not None else None

    def _upsert(self, table_name: str, payload: dict, unique_keys=None):
        cur = self.conn.cursor()
        if unique_keys:
            where = " AND ".join([f"{k}=?" for k in unique_keys])
            existing = cur.execute(
                f"SELECT id FROM {table_name} WHERE {where}",
                [payload[k] for k in unique_keys],
            ).fetchone()
            if existing:
                payload = dict(payload)
                payload["id"] = existing["id"]
        if payload.get("id"):
            row_id = payload["id"]
            cols = [k for k in payload.keys() if k != "id"]
            cur.execute(
                f"UPDATE {table_name} SET " + ", ".join([f"{c}=?" for c in cols]) + " WHERE id=?",
                [payload[c] for c in cols] + [row_id],
            )
        else:
            cols = list(payload.keys())
            cur.execute(
                f"INSERT INTO {table_name} (" + ", ".join(cols) + ") VALUES (" + ", ".join(["?"] * len(cols)) + ")",
                [payload[c] for c in cols],
            )
            row_id = cur.lastrowid
        self.conn.commit()
        return self.get_by_id(table_name, row_id)

    def get_by_id(self, table_name: str, row_id: int):
        row = self.conn.execute(f"SELECT * FROM {table_name} WHERE id=?", (row_id,)).fetchone()
        return self.row_to_dict(row)

    def upsert_company(self, payload: dict):
        payload = dict(payload)
        payload["updated_at"] = datetime.now().isoformat(timespec="seconds")
        if not payload.get("created_at"):
            payload.setdefault("created_at", datetime.now().isoformat(timespec="seconds"))
        return self._upsert("companies", payload)

    def upsert_sheet(self, table_name: str, payload: dict):
        payload = dict(payload)
        payload["updated_at"] = datetime.now().isoformat(timespec="seconds")
        return self._upsert(table_name, payload, unique_keys=["company_id", "year"])

    def upsert_ratio_note(self, company_id: int, ratio_name: str, note: str):
        payload = {
            "company_id": company_id,
            "ratio_name": ratio_name,
            "note": note,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        return self._upsert("ratio_notes", payload, unique_keys=["company_id", "ratio_name"])

    def get_ratio_note(self, company_id: int, ratio_name: str) -> str:
        row = self.conn.execute(
            "SELECT note FROM ratio_notes WHERE company_id=? AND ratio_name=?",
            (company_id, ratio_name)
        ).fetchone()
        return row["note"] if row else ""

    def delete_company(self, company_id: int):
        self.conn.execute("DELETE FROM companies WHERE id=?", (company_id,))
        self.conn.commit()

    def get_company(self, company_id: int):
        return self.get_by_id("companies", company_id)

    def search_companies(self, term: str = ""):
        term = f"%{term.strip().lower()}%"
        rows = self.conn.execute(
            """
            SELECT * FROM companies
            WHERE ?='%%' OR lower(name) LIKE ? OR lower(afm) LIKE ? OR lower(gemi) LIKE ?
            ORDER BY lower(name)
            """,
            (term, term, term, term),
        ).fetchall()
        return [self.row_to_dict(r) for r in rows]

    def search_sheets(self, table_name: str, term: str = ""):
        term = f"%{term.strip().lower()}%"
        rows = self.conn.execute(
            f"""
            SELECT s.*, c.name AS company_name, c.afm AS company_afm, c.gemi AS company_gemi
            FROM {table_name} s
            JOIN companies c ON c.id = s.company_id
            WHERE ?='%%' OR lower(c.name) LIKE ? OR lower(c.afm) LIKE ? OR lower(c.gemi) LIKE ? OR cast(s.year as text) LIKE ?
            ORDER BY lower(c.name), s.year DESC
            """,
            (term, term, term, term, term),
        ).fetchall()
        results = []
        for r in rows:
            d = self.row_to_dict(r)
            comp = {"id": d["company_id"], "name": d["company_name"], "afm": d["company_afm"], "gemi": d["company_gemi"]}
            results.append((comp, d))
        return results

    def get_sheet(self, table_name: str, company_id: int, year: int):
        row = self.conn.execute(
            f"SELECT * FROM {table_name} WHERE company_id=? AND year=?",
            (company_id, year),
        ).fetchone()
        return self.row_to_dict(row)

    def years_for_company(self, company_id: int):
        rows = self.conn.execute(
            """
            SELECT year FROM balance_sheets WHERE company_id=?
            UNION
            SELECT year FROM income_sheets WHERE company_id=?
            ORDER BY year DESC
            """,
            (company_id, company_id),
        ).fetchall()
        return [int(r[0]) for r in rows]

    def all_years(self):
        rows = self.conn.execute(
            """
            SELECT year FROM balance_sheets
            UNION
            SELECT year FROM income_sheets
            ORDER BY year DESC
            """
        ).fetchall()
        return [int(r[0]) for r in rows]

    def all_companies_with_data(self):
        rows = self.conn.execute(
            """
            SELECT DISTINCT c.*
            FROM companies c
            JOIN balance_sheets b ON b.company_id = c.id
            JOIN income_sheets i ON i.company_id = c.id AND i.year = b.year
            ORDER BY lower(c.name)
            """
        ).fetchall()
        return [self.row_to_dict(r) for r in rows]


def calc_balance_totals(data: dict) -> dict:
    non_current = sum(safe_float(data.get(k)) for k in [
        "fixed_tangible", "intangible_assets", "subsidiary_investments", "other_company_investments",
        "right_of_use", "investment_property", "other_long_term_receivables", "non_current_diff"
    ])
    current = sum(safe_float(data.get(k)) for k in [
        "inventory", "trade_receivables", "other_short_term_receivables",
        "other_current_assets", "cash_and_equivalents", "current_diff"
    ])
    total_assets = non_current + current
    equity = sum(safe_float(data.get(k)) for k in [
        "share_capital", "share_premium", "reserves", "retained_earnings", "equity_diff"
    ])
    liabilities = safe_float(data.get("long_term_liabilities")) + safe_float(data.get("short_term_liabilities")) + safe_float(data.get("liabilities_diff"))
    total_liabilities_equity = equity + liabilities
    return {
        "non_current_assets": non_current,
        "current_assets": current,
        "total_assets": total_assets,
        "equity": equity,
        "liabilities": liabilities,
        "total_liabilities_equity": total_liabilities_equity,
    }


def signed_income_amount(value) -> float:
    """Income/revenue rows should contribute positively; keeps genuinely negative other results."""
    return safe_float(value)


def expense_amount(value) -> float:
    """Expense/cost/tax/depreciation rows are treated as costs regardless of stored sign."""
    return abs(safe_float(value))


def calc_income_totals(data: dict) -> dict:
    """
    Income Statement subtotals aligned to the latest Excel template.

    The app now displays expense rows as positive amounts. Therefore totals subtract
    expense fields and add income fields explicitly, instead of relying on old
    negative-value summing conventions.
    """
    net_sales = safe_float(data.get("net_sales"))
    state_grants = safe_float(data.get("state_grants"))
    cost_of_sales = expense_amount(data.get("cost_of_sales"))

    gross_profit = net_sales + state_grants - cost_of_sales

    # Operating expenses
    admin_expenses = expense_amount(data.get("admin_expenses"))
    selling_expenses = expense_amount(data.get("selling_expenses"))
    research_expenses = expense_amount(data.get("research_expenses"))
    provisions_expenses = expense_amount(data.get("provisions_expenses"))
    other_operating_expenses = expense_amount(data.get("other_operating_expenses"))
    inventory_revaluation_losses = expense_amount(data.get("inventory_revaluation_losses"))
    other_expenses = expense_amount(data.get("other_expenses"))

    # Operating income
    operating_income = safe_float(data.get("other_operating_income"))
    inventory_revaluation_gains = safe_float(data.get("inventory_revaluation_gains"))
    other_income = safe_float(data.get("other_income"))

    ebit = (
        gross_profit
        + operating_income
        + inventory_revaluation_gains
        + other_income
        - admin_expenses
        - selling_expenses
        - research_expenses
        - provisions_expenses
        - other_operating_expenses
        - inventory_revaluation_losses
        - other_expenses
    )

    depr_tangible = expense_amount(data.get("depr_tangible"))
    depr_intangible = expense_amount(data.get("depr_intangible"))
    depreciation = depr_tangible + depr_intangible

    ebitda = safe_float(data.get("ebitda_manual"))

    # Financial block: dividends/income/gains add; expenses/losses subtract.
    dividend_income = safe_float(data.get("dividend_income"))
    financial_expenses = expense_amount(data.get("financial_expenses"))
    other_financial_losses = expense_amount(data.get("other_financial_losses"))
    financial_income = safe_float(data.get("financial_income"))
    other_financial_gains = safe_float(data.get("other_financial_gains"))

    pbt = (
        ebit
        + dividend_income
        + financial_income
        + other_financial_gains
        - financial_expenses
        - other_financial_losses
    )

    # Taxes may be negative in the template, so keep their sign.
    income_taxes = safe_float(data.get("income_taxes"))
    pat = pbt - income_taxes

    total_revenue_for_turnover = net_sales + operating_income

    return {
        "gross_profit": gross_profit,
        "ebit": ebit,
        "ebitda": ebitda,
        "depreciation": depreciation,
        "pbt": pbt,
        "pat": pat,
        "total_revenue_for_turnover": total_revenue_for_turnover,
    }

def build_analysis(company: dict, year: int, balance: dict, income: dict, prev_balance: dict | None = None):
    bt = calc_balance_totals(balance)
    it = calc_income_totals(income)
    prev_bt = calc_balance_totals(prev_balance) if prev_balance else None

    # Investment ratios 
    profit_margin = div(it["pat"], safe_float(income.get("net_sales")))
    asset_turnover = div(safe_float(income.get("net_sales")), bt["total_assets"])
    roa = None if profit_margin is None or asset_turnover is None else profit_margin * asset_turnover
    financial_leverage = div(bt["total_liabilities_equity"], bt["equity"])
    roe = None if roa is None or financial_leverage is None else roa * financial_leverage

    # Liquidity ratios 
    current_ratio = div(bt["current_assets"], safe_float(balance.get("short_term_liabilities")))
    quick_ratio = div(bt["current_assets"] - safe_float(balance.get("inventory")), safe_float(balance.get("short_term_liabilities")))

    inventory_turnover = receivable_turnover = payable_turnover = operating_cycle = None
    inventory_days = receivable_days = payable_days = None
    fixed_asset_turnover = total_asset_turnover_2 = None

    working_capital_requirements = bt["current_assets"] - safe_float(balance.get("short_term_liabilities"))

    if prev_balance:
        avg_inventory = (safe_float(balance.get("inventory")) + safe_float(prev_balance.get("inventory"))) / 2
        avg_receivables = (safe_float(balance.get("trade_receivables")) + safe_float(prev_balance.get("trade_receivables"))) / 2
        avg_payables = (safe_float(balance.get("short_term_liabilities")) + safe_float(prev_balance.get("short_term_liabilities"))) / 2

        # The app stores / displays costs as positive amounts, but the template uses
        # a leading minus in these turnover formulas.  Using -abs(...) makes the
        # result stable even if an older database still contains negative expense rows.
        cost_for_turnover = expense_amount(income.get("cost_of_sales"))
        signed_cost_for_turnover = -cost_for_turnover

        inventory_turnover = div(signed_cost_for_turnover, avg_inventory)
        inventory_days = div(365, inventory_turnover) if inventory_turnover not in (None, 0) else None

        receivable_turnover = div(signed_cost_for_turnover, avg_receivables)
        receivable_days = div(365, receivable_turnover) if receivable_turnover not in (None, 0) else None

        inventory_change = safe_float(balance.get("inventory")) - safe_float(prev_balance.get("inventory"))
        payable_turnover = div(-(cost_for_turnover + inventory_change), avg_payables)
        payable_days = div(365, payable_turnover) if payable_turnover not in (None, 0) else None

        if payable_days is not None and receivable_days is not None:
            operating_cycle = receivable_days - payable_days

        prev_bt_for_turnover = calc_balance_totals(prev_balance)
        total_revenue_for_turnover = safe_float(income.get("net_sales")) + safe_float(income.get("other_operating_income"))
        fixed_asset_turnover = div(total_revenue_for_turnover, prev_bt_for_turnover["non_current_assets"] - bt["non_current_assets"])
        total_asset_turnover_2 = div(total_revenue_for_turnover, prev_bt_for_turnover["total_assets"] - bt["total_assets"])

    # The updated template's EBIT in the indicators area is EBITDA - Depreciations.
    indicator_ebit = it["ebitda"] - it["depreciation"]

    return {
        "company_name": company.get("name", ""),
        "year": year,
        "totals": {**bt, **it, "working_capital_requirements": working_capital_requirements, "indicator_ebit": indicator_ebit},
        "ratios": {
            "Profit Margin (Περιθώριο Καθαρού Κέρδους)": profit_margin,
            "Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)": asset_turnover,
            "ROA (Απόδοση Συνόλου Ενεργητικού)": roa,
            "Financial Leverage (Χρημ/ική Μόχλευση)": financial_leverage,
            "ROE (Απόδοση Ι/Κ)": roe,
            "Current Ratio (Κεφ./Κίνησης)": current_ratio,
            "Quick Ratio (Ρευστότητας)": quick_ratio,
            "Inventory Turnover Ratio (Κυκλοφορ. Ταχύτ. Αποθεμάτων σε φορές)": inventory_turnover,
            "Inventory Days (Παραμονή Αποθεμάτων σε ημέρες)": inventory_days,
            "Receivable Turnover Ratio (Κυκλοφορ. Ταχύτ. Απαιτήσεων σε φορές)": receivable_turnover,
            "Receivable Days (Παραμονή Απαιτήσεων σε ημέρες)": receivable_days,
            "Payable Turnover Ratio (Κυκλοφορ. Ταχύτ. Υποχρεώσεων σε φορές)": payable_turnover,
            "Payable Days (Παραμονή Βραχυπρ. Υποχρεώσεων σε ημέρες)": payable_days,
            "Operating Cycle (Λειτουργικός Κύκλος σε ημέρες)": operating_cycle,
            "Working Capital Requirements (Ανάγκες σε Κ/Κ)": working_capital_requirements,
            "Fixed Asset Turnover (Κυκλοφορ. Ταχύτ. Πάγιου Ενεργητικού σε φορές)": fixed_asset_turnover,
            "Total Asset Turnover (Κυκλοφορ. Ταχύτ. Ενεργητικού σε φορές)": total_asset_turnover_2,
            "EBIT": indicator_ebit,
            "EBITDA": it["ebitda"],
            "Depreciations (Αποσβέσεις Χρήσης)": it["depreciation"],
        },
    }

def extract_years_from_balance_sheet(ws):
    years = {}
    for col in range(2, 6):
        val = ws.cell(2, col).value
        if val:
            token = str(val).strip().split()[0]
            if token.isdigit():
                years[col] = int(token)
    return years


def extract_years_from_income_sheet(ws):
    years = {}
    for col in [2, 4, 6]:
        val = ws.cell(1, col).value
        if val:
            token = str(val).strip().split()[0]
            if token.isdigit():
                years[col] = int(token)
    return years



def clean_import_value(value):
    """Import helper: workbook placeholder 0.001 values are treated as real zero."""
    v = safe_float(value)
    return 0.0 if abs(v) < 0.01 else v



def _formula_cell_value(ws_values, ws_formulas, row: int, col: int):
    """Return an import cell value, with fallback evaluation for simple formulas.

    Some template input cells may contain formulas (for example =11584136+1959091).
    When a workbook has not been recalculated by Excel, openpyxl(data_only=True)
    may return None for those cells. This helper uses the formula text as a fallback
    and evaluates only safe, simple arithmetic / SUM formulas needed by the import template.
    """
    value = ws_values.cell(row, col).value
    if value not in (None, ""):
        return value

    formula = ws_formulas.cell(row, col).value
    if not isinstance(formula, str) or not formula.startswith("="):
        return value

    expr = formula[1:].strip()
    if not expr:
        return value

    import re

    def cell_value(ref: str) -> float:
        try:
            v = ws_values[ref].value
            if v in (None, ""):
                raw = ws_formulas[ref].value
                if isinstance(raw, str) and raw.startswith("="):
                    return float(_formula_cell_value(ws_values, ws_formulas, ws_formulas[ref].row, ws_formulas[ref].column) or 0)
                v = raw
            return safe_float(v)
        except Exception:
            return 0.0

    def eval_sum(match):
        range_text = match.group(1)
        total = 0.0
        try:
            for row_cells in ws_formulas[range_text]:
                for c in row_cells:
                    total += cell_value(c.coordinate)
        except Exception:
            return "0"
        return str(total)

    expr = re.sub(r"SUM\(([^)]+)\)", eval_sum, expr, flags=re.IGNORECASE)

    # Replace remaining cell references, e.g. B3 or $B$3.
    expr = re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", lambda m: str(cell_value(f"{m.group(1)}{m.group(2)}")), expr)

    # Evaluate only plain numeric arithmetic.
    if not re.fullmatch(r"[0-9\.\,\+\-\*/\(\)\s]+", expr):
        return value
    try:
        return eval(expr.replace(",", "."), {"__builtins__": {}}, {})
    except Exception:
        return value

def parse_excel_template(file_path: str):
    if not OPENPYXL_OK:
        raise RuntimeError("Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη.")

    # Open twice: values for normal imports, formulas as fallback when Excel has not cached formula results.
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)

    required = {"Κατάσταση Οικονομικής Θέσης", "Κατάσταση Συνολικού Εισοδήματος"}
    if not required.issubset(set(wb_values.sheetnames)):
        raise RuntimeError("Το αρχείο Excel δεν έχει τα απαιτούμενα sheets import.")

    balance_ws = wb_values["Κατάσταση Οικονομικής Θέσης"]
    income_ws = wb_values["Κατάσταση Συνολικού Εισοδήματος"]
    balance_ws_f = wb_formulas["Κατάσταση Οικονομικής Θέσης"]
    income_ws_f = wb_formulas["Κατάσταση Συνολικού Εισοδήματος"]

    balance_years = extract_years_from_balance_sheet(balance_ws)
    income_years = extract_years_from_income_sheet(income_ws)

    parsed = {}
    for col, year in balance_years.items():
        parsed.setdefault(year, {"balance": {}, "income": {}})
        for row, key in BALANCE_IMPORT_MAP.items():
            value = _formula_cell_value(balance_ws, balance_ws_f, row, col)
            parsed[year]["balance"][key] = clean_import_value(value)

    expense_like = {
        "cost_of_sales", "admin_expenses", "selling_expenses", "research_expenses",
        "provisions_expenses", "other_operating_expenses", "inventory_revaluation_losses",
        "other_expenses", "depr_tangible", "depr_intangible", "financial_expenses",
        "other_financial_losses"
    }

    for col, year in income_years.items():
        parsed.setdefault(year, {"balance": {}, "income": {}})
        for row, key in INCOME_IMPORT_MAP.items():
            value = clean_import_value(_formula_cell_value(income_ws, income_ws_f, row, col))
            if key in expense_like:
                value = abs(value)
            parsed[year]["income"][key] = value

    return {year: values for year, values in parsed.items() if values["balance"] or values["income"]}



def apply_template_third_year_adjustment(analyses: list[dict], company: dict, db: SQLiteDB) -> list[dict]:
    """Match the latest CorrectCalcs.xlsx report layout for the oldest displayed year.

    The report shows the latest 3 years. In the reference workbook, the 3rd displayed
    year uses the previous balance-sheet column for several investment / liquidity
    ratios (while the income statement remains the 3rd-year income statement).
    This fixes the wrong 3rd-column values seen in the PDF export.
    """
    if len(analyses) < 3:
        return analyses
    oldest = analyses[2]
    year = int(oldest.get("year"))
    prev_balance = db.get_sheet("balance_sheets", company["id"], year - 1)
    income = db.get_sheet("income_sheets", company["id"], year)
    if not prev_balance or not income:
        return analyses

    bt_prev = calc_balance_totals(prev_balance)
    it = oldest.get("totals", {})
    net_sales = safe_float(income.get("net_sales"))

    profit_margin = div(it.get("pat"), net_sales)
    asset_turnover = div(net_sales, bt_prev["total_assets"])
    roa = None if profit_margin is None or asset_turnover is None else profit_margin * asset_turnover
    financial_leverage = div(bt_prev["total_liabilities_equity"], bt_prev["equity"])
    roe = None if roa is None or financial_leverage is None else roa * financial_leverage

    ratios = oldest["ratios"]
    ratios["Profit Margin (Περιθώριο Καθαρού Κέρδους)"] = profit_margin
    ratios["Asset Turnover (Κεφαλ. Παραγωγικ. Ενεργητικού)"] = asset_turnover
    ratios["ROA (Απόδοση Συνόλου Ενεργητικού)"] = roa
    ratios["Financial Leverage (Χρημ/ική Μόχλευση)"] = financial_leverage
    ratios["ROE (Απόδοση Ι/Κ)"] = roe
    ratios["Current Ratio (Κεφ./Κίνησης)"] = div(bt_prev["current_assets"], safe_float(prev_balance.get("short_term_liabilities")))
    return analyses


class CompanyTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.current_id = None
        self.vars = {
            "name": tk.StringVar(),
            "afm": tk.StringVar(),
            "gemi": tk.StringVar(),
            "gemi_url": tk.StringVar(),
            "company_url": tk.StringVar(),
        }
        self.build_ui()
        self.refresh_list()

    def build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text="Αναζήτηση:").pack(side="left")
        self.search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.search_var, width=40).pack(side="left", padx=5)
        ttk.Button(top, text="Search", command=self.refresh_list).pack(side="left")
        ttk.Button(top, text="New", command=self.new_form).pack(side="left", padx=(15, 5))
        ttk.Button(top, text="Save", command=self.save_company).pack(side="left", padx=5)
        ttk.Button(top, text="Edit selected", command=self.load_selected).pack(side="left", padx=5)
        ttk.Button(top, text="Delete", command=self.delete_selected).pack(side="left", padx=5)
        ttk.Button(top, text="Import Excel", command=self.import_excel).pack(side="left", padx=(15, 5))

        content = ttk.Frame(self)
        content.pack(fill="both", expand=True, padx=10, pady=5)
        left = ttk.LabelFrame(content, text="Companies")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right = ttk.LabelFrame(content, text="Company data")
        right.pack(side="left", fill="both", expand=True)

        self.tree = ttk.Treeview(left, columns=("name", "afm", "gemi"), show="headings", height=18)
        for col, title, width in [("name", "Επωνυμία", 220), ("afm", "ΑΦΜ", 110), ("gemi", "ΓΕΜΗ", 110)]:
            self.tree.heading(col, text=title)
            self.tree.column(col, width=width, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<Double-1>", lambda e: self.load_selected())


        form = ttk.Frame(right)
        form.pack(fill="both", expand=True, padx=10, pady=10)
        simple_fields = COMPANY_FIELDS[:-1]
        for i, (key, label) in enumerate(simple_fields):
            ttk.Label(form, text=label).grid(row=i, column=0, sticky="w", pady=4, padx=(0, 8))
            entry = ttk.Entry(form, textvariable=self.vars[key], width=55)
            entry.grid(row=i, column=1, sticky="ew", pady=4)
            if key in {"gemi_url", "company_url"}:
                ttk.Button(
                    form,
                    text="Open",
                    command=lambda k=key: self.open_company_url(k)
                ).grid(row=i, column=2, sticky="w", padx=(6, 0), pady=4)
        notes_row = len(simple_fields)
        ttk.Label(form, text="Σημειώσεις").grid(row=notes_row, column=0, sticky="nw", pady=4, padx=(0, 8))
        self.notes_text = tk.Text(form, height=10, width=60, wrap="word")
        self.notes_text.grid(row=notes_row, column=1, columnspan=2, sticky="nsew", pady=4)
        form.columnconfigure(1, weight=1)
        form.rowconfigure(notes_row, weight=1)

    def open_company_url(self, key: str):
        try:
            open_url(self.vars[key].get())
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def refresh_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.app.db.search_companies(self.search_var.get()):
            self.tree.insert("", "end", iid=str(row["id"]), values=(row.get("name", ""), row.get("afm", ""), row.get("gemi", "")))
        self.app.refresh_company_combos()

    def new_form(self):
        self.current_id = None
        for v in self.vars.values():
            v.set("")
        self.notes_text.delete("1.0", "end")

    def load_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Επίλεξε πρώτα εταιρεία από τη λίστα.")
            return
        row = self.app.db.get_company(int(sel[0]))
        if not row:
            return
        self.current_id = row["id"]
        for key in self.vars:
            self.vars[key].set(row.get(key, ""))
        self.notes_text.delete("1.0", "end")
        self.notes_text.insert("1.0", row.get("notes", ""))

    def save_company(self):
        if not self.vars["name"].get().strip():
            messagebox.showerror(APP_TITLE, "Η Επωνυμία είναι υποχρεωτική.")
            return
        payload = {key: self.vars[key].get().strip() for key in self.vars}
        payload["notes"] = self.notes_text.get("1.0", "end").strip()
        if self.current_id:
            payload["id"] = self.current_id
        saved = self.app.db.upsert_company(payload)
        self.current_id = saved["id"]
        self.refresh_list()
        messagebox.showinfo(APP_TITLE, "Η εταιρεία αποθηκεύτηκε.")

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Επίλεξε πρώτα εταιρεία από τη λίστα.")
            return
        row = self.app.db.get_company(int(sel[0]))
        if not row:
            return
        answer = messagebox.askyesno(
            APP_TITLE,
            f"Να διαγραφεί η εταιρεία '{row.get('name', '')}' και όλα τα σχετικά δεδομένα της;"
        )
        if not answer:
            return
        self.app.db.delete_company(int(sel[0]))
        self.new_form()
        self.refresh_list()
        if hasattr(self.app, "balance_tab"):
            self.app.balance_tab.refresh_records()
        if hasattr(self.app, "income_tab"):
            self.app.income_tab.refresh_records()
        if hasattr(self.app, "analysis_tab"):
            self.app.analysis_tab.clear_report()
        if hasattr(self.app, "compare_tab"):
            self.app.compare_tab.refresh_year_options()
            self.app.compare_tab.clear_results()
        messagebox.showinfo(APP_TITLE, "Η εταιρεία διαγράφηκε.")

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")], title="Επίλεξε αρχείο Import Template")
        if not file_path:
            return
        try:
            company_id = self.app.ask_company_for_import()
            parsed = parse_excel_template(file_path)
            if not parsed:
                raise RuntimeError("Δεν βρέθηκαν οικονομικά δεδομένα για import.")
            imported_years = []
            for year, content in parsed.items():
                bal = {k: v for k, v in content.get("balance", {}).items()}
                inc = {k: v for k, v in content.get("income", {}).items()}
                if bal:
                    bal.setdefault("comments", "")
                    self.app.db.upsert_sheet("balance_sheets", {"company_id": company_id, "year": year, **bal})
                if inc:
                    inc.setdefault("comments", "")
                    self.app.db.upsert_sheet("income_sheets", {"company_id": company_id, "year": year, **inc})
                imported_years.append(str(year))
            self.refresh_list()
            self.app.balance_tab.refresh_records()
            self.app.income_tab.refresh_records()
            self.app.refresh_company_combos()
            messagebox.showinfo(APP_TITLE, f"Το import ολοκληρώθηκε για τα έτη: {', '.join(sorted(imported_years, reverse=True))}")
        except ImportCanceled:
            return
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία import Excel:\n{e}")


class ScrollableEntryPanel(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.v_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.v_scrollbar.pack(side="right", fill="y")
        self._bind_mousewheel(self.canvas)
        self._bind_mousewheel(self.inner)

    def _on_inner_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.window_id, width=event.width)

    def _on_mousewheel(self, event):
        if event.delta:
            self.canvas.yview_scroll(int(-event.delta / 120), "units")
        elif getattr(event, "num", None) == 4:
            self.canvas.yview_scroll(-3, "units")
        elif getattr(event, "num", None) == 5:
            self.canvas.yview_scroll(3, "units")
        return "break"

    def _bind_mousewheel(self, widget):
        widget.bind("<MouseWheel>", self._on_mousewheel, add="+")
        widget.bind("<Button-4>", self._on_mousewheel, add="+")
        widget.bind("<Button-5>", self._on_mousewheel, add="+")


class BaseStatementTab(ttk.Frame):
    def __init__(self, master, app, title: str, table_name: str):
        super().__init__(master)
        self.app = app
        self.title = title
        self.table_name = table_name
        self.company_map = {}
        self.build_shell()
        self.refresh_records()

    def build_shell(self):
        controls = ttk.Frame(self)
        controls.pack(fill="x", padx=10, pady=8)

        ttk.Label(controls, text="Εταιρεία:").grid(row=0, column=0, sticky="w")
        self.company_var = tk.StringVar()
        self.company_combo = ttk.Combobox(controls, textvariable=self.company_var, state="readonly", width=40)
        self.company_combo.grid(row=0, column=1, sticky="w", padx=5)

        ttk.Label(controls, text="Έτος:").grid(row=0, column=2, sticky="w", padx=(10, 0))
        self.year_var = tk.StringVar()
        ttk.Entry(controls, textvariable=self.year_var, width=10).grid(row=0, column=3, sticky="w", padx=5)

        ttk.Label(controls, text="Αναζήτηση:").grid(row=0, column=4, sticky="w", padx=(15, 0))
        self.search_var = tk.StringVar()
        ttk.Entry(controls, textvariable=self.search_var, width=30).grid(row=0, column=5, sticky="w", padx=5)
        ttk.Button(controls, text="Search", command=self.refresh_records).grid(row=0, column=6, padx=5)
        ttk.Button(controls, text="New sheet", command=self.new_form).grid(row=0, column=7, padx=(15, 5))
        ttk.Button(controls, text="Save sheet", command=self.save_sheet).grid(row=0, column=8, padx=5)
        ttk.Button(controls, text="Edit sheet", command=self.load_selected).grid(row=0, column=9, padx=5)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True, padx=10, pady=5)
        left = ttk.LabelFrame(main, text="Saved sheets")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right = ttk.LabelFrame(main, text=self.title, width=680)
        right.pack(side="left", fill="both", expand=True)
        right.pack_propagate(False)

        self.tree = ttk.Treeview(left, columns=("company", "afm", "gemi", "year"), show="headings", height=18)
        for col, title, width in [("company", "Επωνυμία", 180), ("afm", "ΑΦΜ", 90), ("gemi", "ΓΕΜΗ", 95), ("year", "Έτος", 60)]:
            self.tree.heading(col, text=title)
            self.tree.column(col, width=width, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<Double-1>", lambda e: self.load_selected())

        self.panel = ScrollableEntryPanel(right)
        self.panel.pack(fill="both", expand=True)

    def set_company_options(self, options):
        self.company_map = {text: cid for cid, text in options}
        self.company_combo["values"] = [text for _, text in options]

    def refresh_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for comp, row in self.app.db.search_sheets(self.table_name, self.search_var.get()):
            iid = f"{row['company_id']}|{row['year']}"
            self.tree.insert("", "end", iid=iid, values=(comp.get("name", ""), comp.get("afm", ""), comp.get("gemi", ""), row.get("year", "")))

    def new_form(self):
        self.company_var.set("")
        self.year_var.set("")
        self.reset_fields()

    def load_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Επίλεξε πρώτα εγγραφή από τη λίστα.")
            return
        company_id, year = sel[0].split("|")
        row = self.app.db.get_sheet(self.table_name, int(company_id), int(year))
        comp = self.app.db.get_company(int(company_id))
        if not row or not comp:
            return
        self.company_var.set(self.app.company_display(comp))
        self.year_var.set(str(year))
        self.populate_fields(row)

    def save_sheet(self):
        company_text = self.company_var.get().strip()
        if company_text not in self.company_map:
            messagebox.showerror(APP_TITLE, "Επίλεξε εταιρεία.")
            return
        if not self.year_var.get().strip().isdigit():
            messagebox.showerror(APP_TITLE, "Το έτος πρέπει να είναι ακέραιος αριθμός.")
            return
        payload = {
            "company_id": self.company_map[company_text],
            "year": int(self.year_var.get().strip()),
        }
        payload.update(self.collect_payload())
        self.app.db.upsert_sheet(self.table_name, payload)
        self.refresh_records()
        messagebox.showinfo(APP_TITLE, "Το φύλλο αποθηκεύτηκε.")

    def reset_fields(self):
        raise NotImplementedError

    def populate_fields(self, row):
        raise NotImplementedError

    def collect_payload(self):
        raise NotImplementedError


class BalanceSheetTab(BaseStatementTab):
    def __init__(self, master, app):
        self.input_vars = {k: tk.StringVar() for k in BALANCE_NUMERIC_FIELDS}
        self.auto_vars = {k: tk.StringVar() for k in ["non_current_assets", "current_assets", "total_assets", "equity", "liabilities", "total_liabilities_equity"]}
        super().__init__(master, app, "Ισολογισμός", "balance_sheets")
        self.build_form()

    def build_form(self):
        inner = self.panel.inner
        self.entry_widgets = {}
        row = 0
        for kind, label, key in BALANCE_LAYOUT:
            if kind == "space":
                ttk.Label(inner, text="").grid(row=row, column=0, columnspan=2, pady=4)
            elif kind == "header":
                ttk.Label(inner, text=label, font=("Segoe UI", 12, "bold")).grid(row=row, column=0, columnspan=2, sticky="ew", pady=(8, 4))
            elif kind == "field" or kind == "diff":
                ttk.Label(inner, text=label, wraplength=500, justify="left").grid(row=row, column=0, sticky="w", padx=(8, 8), pady=3)
                e = ttk.Entry(inner, textvariable=self.input_vars[key], width=18)
                e.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=3)
                e.bind("<KeyRelease>", lambda event: self.update_auto_fields())
                e.bind("<FocusIn>", lambda event, ent=e: unformat_entry_thousands(ent))
                e.bind("<FocusOut>", lambda event, ent=e: (format_entry_thousands(ent), self.update_auto_fields()))
                e.bind("<FocusIn>", lambda event, ent=e: unformat_entry_thousands(ent))
                e.bind("<FocusOut>", lambda event, ent=e: (format_entry_thousands(ent), self.update_auto_fields()))
                self.entry_widgets[key] = e
            elif kind == "auto":
                ttk.Label(inner, text=label, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="w", padx=(8, 8), pady=3)
                e = ttk.Entry(inner, textvariable=self.auto_vars[key], width=24, state="readonly")
                e.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=3)
            elif kind == "auto_header":
                ttk.Label(inner, text=label, font=("Segoe UI", 11, "bold")).grid(row=row, column=0, sticky="w", padx=(8, 8), pady=(8, 3))
                e = ttk.Entry(inner, textvariable=self.auto_vars[key], width=24, state="readonly")
                e.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=(8, 3))
            row += 1

        ttk.Label(inner, text="Σχόλια").grid(row=row, column=0, sticky="nw", padx=(8, 8), pady=(10, 4))
        self.comments_text = tk.Text(inner, height=6, width=30, wrap="word")
        self.comments_text.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=(10, 4))
        inner.columnconfigure(1, weight=1)
        self.update_auto_fields()

    def update_auto_fields(self):
        data = {k: safe_float(v.get()) for k, v in self.input_vars.items()}
        totals = calc_balance_totals(data)
        for key, var in self.auto_vars.items():
            var.set(fmt_num(totals.get(key, 0)))

    def reset_fields(self):
        for v in self.input_vars.values():
            v.set("")
        self.comments_text.delete("1.0", "end")
        self.update_auto_fields()

    def populate_fields(self, row):
        for k in self.input_vars:
            self.input_vars[k].set("" if abs(safe_float(row.get(k))) < 0.0000001 else format_thousands_dot(row.get(k, "")))
        self.comments_text.delete("1.0", "end")
        self.comments_text.insert("1.0", row.get("comments", ""))
        self.update_auto_fields()

    def collect_payload(self):
        payload = {k: safe_float(v.get()) for k, v in self.input_vars.items()}
        payload["comments"] = self.comments_text.get("1.0", "end").strip()
        return payload


class IncomeSheetTab(BaseStatementTab):
    def __init__(self, master, app):
        self.input_vars = {k: tk.StringVar() for k in INCOME_NUMERIC_FIELDS}
        self.auto_vars = {k: tk.StringVar() for k in ["gross_profit", "ebit", "depreciation", "pbt", "pat"]}
        super().__init__(master, app, "Αποτελέσματα Χρήσης", "income_sheets")
        self.build_form()

    def build_form(self):
        inner = self.panel.inner
        row = 0
        for kind, label, key in INCOME_LAYOUT:
            if kind == "space":
                ttk.Label(inner, text="").grid(row=row, column=0, columnspan=2, pady=4)
            elif kind == "header":
                ttk.Label(inner, text=label, font=("Segoe UI", 11, "bold")).grid(row=row, column=0, columnspan=2, sticky="ew", padx=(8, 8), pady=(8, 4))
            elif kind in {"field", "field_bold"}:
                font = ("Segoe UI", 10, "bold") if kind == "field_bold" else None
                ttk.Label(inner, text=label, font=font, wraplength=500, justify="left").grid(row=row, column=0, sticky="w", padx=(8, 8), pady=3)
                e = ttk.Entry(inner, textvariable=self.input_vars[key], width=18)
                e.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=3)
                e.bind("<KeyRelease>", lambda event: self.update_auto_fields())
                e.bind("<FocusIn>", lambda event, ent=e: unformat_entry_thousands(ent))
                e.bind("<FocusOut>", lambda event, ent=e: (format_entry_thousands(ent), self.update_auto_fields()))
            elif kind == "auto":
                ttk.Label(inner, text=label, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="w", padx=(8, 8), pady=3)
                ttk.Entry(inner, textvariable=self.auto_vars[key], width=24, state="readonly").grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=3)
            row += 1
        ttk.Label(inner, text="Σχόλια").grid(row=row, column=0, sticky="nw", padx=(8, 8), pady=(10, 4))
        self.comments_text = tk.Text(inner, height=6, width=30, wrap="word")
        self.comments_text.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=(10, 4))
        inner.columnconfigure(1, weight=1)
        self.update_auto_fields()

    def update_auto_fields(self):
        data = {k: safe_float(v.get()) for k, v in self.input_vars.items()}
        totals = calc_income_totals(data)
        for key, var in self.auto_vars.items():
            var.set(fmt_num(totals.get(key, 0)))

    def reset_fields(self):
        for v in self.input_vars.values():
            v.set("")
        self.comments_text.delete("1.0", "end")
        self.update_auto_fields()

    def populate_fields(self, row):
        expense_like = {
            "cost_of_sales", "admin_expenses", "selling_expenses", "research_expenses",
            "provisions_expenses", "other_operating_expenses", "inventory_revaluation_losses",
            "other_expenses", "depr_tangible", "depr_intangible", "financial_expenses",
            "other_financial_losses"
        }
        for k in self.input_vars:
            value = safe_float(row.get(k))
            if k in expense_like:
                value = abs(value)
            self.input_vars[k].set("" if abs(value) < 0.0000001 else format_thousands_dot(value))
        self.comments_text.delete("1.0", "end")
        self.comments_text.insert("1.0", row.get("comments", ""))
        self.update_auto_fields()

    def collect_payload(self):
        if abs(safe_float(self.input_vars["ebitda_manual"].get())) < 0.0000001:
            raise RuntimeError("Το πεδίο EBITDA πρέπει να συμπληρώνεται από τον χρήστη.")
        payload = {k: safe_float(v.get()) for k, v in self.input_vars.items()}
        # v5.3: store expense/cost/depreciation/loss rows as positive values, like the Excel template.
        for k in {
            "cost_of_sales", "admin_expenses", "selling_expenses", "research_expenses",
            "provisions_expenses", "other_operating_expenses", "inventory_revaluation_losses",
            "other_expenses", "depr_tangible", "depr_intangible", "financial_expenses",
            "other_financial_losses"
        }:
            payload[k] = abs(payload.get(k, 0.0))
        payload["comments"] = self.comments_text.get("1.0", "end").strip()
        return payload

    def save_sheet(self):
        try:
            super().save_sheet()
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))



def append_statement_sheet_to_workbook(wb, title: str, company: dict, years: list[int], db: SQLiteDB, statement: str):
    """Append Ισολογισμός / Αποτελέσματα Χρήσης records to the report XLSX."""
    ws = wb.create_sheet(title)
    ws.append(["Εταιρεία", company.get("name", "")])
    ws.append(["ΑΦΜ", company.get("afm", "")])
    ws.append(["ΓΕΜΗ", company.get("gemi", "")])
    ws.append([])

    if statement == "balance":
        table_name = "balance_sheets"
        layout = BALANCE_LAYOUT
        total_func = calc_balance_totals
    else:
        table_name = "income_sheets"
        layout = INCOME_LAYOUT
        total_func = calc_income_totals

    sheets = {year: db.get_sheet(table_name, company["id"], year) or {} for year in years}
    totals = {year: total_func(sheets[year]) for year in years}

    header_row = ws.max_row + 1
    ws.append(["Κονδύλι"] + years)
    style_excel_header(ws, header_row, "D9E2F3")
    group_rows = []

    for kind, label, key in layout:
        if kind == "space":
            ws.append([""] + [None] * len(years))
            continue
        if kind == "header":
            ws.append([label] + [""] * len(years))
            group_rows.append(ws.max_row)
            continue
        if kind in {"field", "diff", "field_bold"}:
            row = [label]
            for year in years:
                row.append(safe_float(sheets[year].get(key)))
            ws.append(row)
            continue
        if kind in {"auto", "auto_header"}:
            row = [label]
            for year in years:
                row.append(totals[year].get(key))
            ws.append(row)
            group_rows.append(ws.max_row)
            continue

    for r in range(header_row + 1, ws.max_row + 1):
        ws.cell(r, 1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        for c in range(2, 2 + len(years)):
            ws.cell(r, c).number_format = '#,##0.00'
    for r in group_rows:
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="E2F0D9")
    ws.freeze_panes = f"B{header_row + 1}"
    autosize_worksheet(ws, min_width=12, max_width=45)

class AnalysisTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.company_map = {}
        self.current_report_rows = []
        self.current_company = None
        self.current_ratio = None
        self.build_ui()

    def build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text="Εταιρεία:").pack(side="left")
        self.company_var = tk.StringVar()
        self.company_combo = ttk.Combobox(top, textvariable=self.company_var, state="readonly", width=40)
        self.company_combo.pack(side="left", padx=5)
        ttk.Button(top, text="Υπολογισμός", command=self.run_analysis).pack(side="left", padx=(10, 5))
        ttk.Button(top, text="Export PDF", command=self.export_pdf).pack(side="left", padx=5)
        ttk.Button(top, text="Export XLSX", command=self.export_xlsx).pack(side="left", padx=5)

        self.info_label = ttk.Label(self, text="Διάλεξε εταιρεία και πάτησε Υπολογισμός.")
        self.info_label.pack(anchor="w", padx=12)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True, padx=10, pady=10)

        left = ttk.Frame(main)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right = ttk.LabelFrame(main, text="Σημείωση δείκτη")
        right.pack(side="left", fill="both", expand=False)

        self.tree = ttk.Treeview(left, columns=("ratio", "y1", "y2", "y3", "y4"), show="headings", height=24)
        self.tree.pack(fill="both", expand=True)
        self.tree.heading("ratio", text="Δείκτης")
        self.tree.column("ratio", width=360, anchor="w")
        for col in ["y1", "y2", "y3", "y4"]:
            self.tree.heading(col, text="-")
            self.tree.column(col, width=115, anchor="center")
        self.tree.bind("<<TreeviewSelect>>", self.on_ratio_selected)

        ttk.Label(right, text="Δείκτης").pack(anchor="w", padx=8, pady=(8, 2))
        self.ratio_name_var = tk.StringVar()
        ttk.Entry(right, textvariable=self.ratio_name_var, state="readonly", width=42).pack(fill="x", padx=8)

        ttk.Label(right, text="Σημείωση").pack(anchor="w", padx=8, pady=(8, 2))
        self.note_text = tk.Text(right, width=40, height=20, wrap="word")
        self.note_text.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        ttk.Button(right, text="Save note", command=self.save_ratio_note).pack(anchor="e", padx=8, pady=(0, 8))

    def set_company_options(self, options):
        self.company_map = {text: cid for cid, text in options}
        self.company_combo["values"] = [text for _, text in options]

    def clear_report(self):
        self.current_report_rows = []
        self.current_company = None
        self.current_ratio = None
        self.company_var.set("")
        self.ratio_name_var.set("")
        self.note_text.delete("1.0", "end")
        self.info_label.configure(text="Διάλεξε εταιρεία και πάτησε Υπολογισμός.")
        for item in self.tree.get_children():
            self.tree.delete(item)
        for col in ["y1", "y2", "y3", "y4"]:
            self.tree.heading(col, text="-")

    def collect_company_analysis(self, company_id: int):
        company = self.app.db.get_company(company_id)
        years = self.app.db.years_for_company(company_id)
        analyses = []
        for year in sorted(years, reverse=True):
            bs = self.app.db.get_sheet("balance_sheets", company_id, year)
            inc = self.app.db.get_sheet("income_sheets", company_id, year)
            if not bs or not inc:
                continue
            prev_bs = self.app.db.get_sheet("balance_sheets", company_id, year - 1)
            analyses.append(build_analysis(company, year, bs, inc, prev_bs if prev_bs else None))
        analyses = apply_template_third_year_adjustment(analyses, company, self.app.db)
        return company, years, analyses

    def run_analysis(self):
        self._save_current_ratio_note_silent()
        company_text = self.company_var.get().strip()
        if company_text not in self.company_map:
            messagebox.showerror(APP_TITLE, "Επίλεξε εταιρεία.")
            return
        company_id = self.company_map[company_text]
        company, years, analyses = self.collect_company_analysis(company_id)
        if not analyses:
            messagebox.showinfo(APP_TITLE, "Χρειάζονται και Ισολογισμός και Αποτελέσματα Χρήσης για το ίδιο έτος.")
            return

        self.current_company = company
        self.current_report_rows = analyses
        visible_years = [a["year"] for a in analyses[:3]]
        for i, col in enumerate(["y1", "y2", "y3", "y4"]):
            self.tree.heading(col, text=str(visible_years[i]) if i < len(visible_years) else "-")

        for item in self.tree.get_children():
            self.tree.delete(item)

        for item_type, label in iter_grouped_ratios():
            if item_type == "__GROUP__":
                row = [label, "", "", "", ""]
                self.tree.insert("", "end", iid=f"group::{label}", values=row, tags=("group_heading",))
                continue
            ratio_name = label
            row = [ratio_name]
            for analysis in analyses[:3]:
                val = analysis["ratios"].get(ratio_name)
                row.append(format_ratio_display(ratio_name, val))
            while len(row) < 5:
                row.append("")
            self.tree.insert("", "end", iid=ratio_name, values=row)

        self.tree.tag_configure("group_heading", font=("Segoe UI", 10, "bold"))
        self.info_label.configure(text=f"Εταιρεία: {company.get('name','')} | Διαθέσιμα έτη: {', '.join(map(str, years))}")
        self.current_ratio = None
        self.ratio_name_var.set("")
        self.note_text.delete("1.0", "end")

    def _save_current_ratio_note_silent(self):
        if not self.current_company or not self.current_ratio:
            return
        note = self.note_text.get("1.0", "end").strip()
        self.app.db.upsert_ratio_note(self.current_company["id"], self.current_ratio, note)

    def on_ratio_selected(self, event=None):
        sel = self.tree.selection()
        if not sel or not self.current_company:
            return
        previous_ratio = self.current_ratio
        new_ratio = sel[0]
        if new_ratio.startswith("group::"):
            return

        if previous_ratio and previous_ratio != new_ratio:
            self._save_current_ratio_note_silent()

        self.current_ratio = new_ratio
        self.ratio_name_var.set(new_ratio)
        note = self.app.db.get_ratio_note(self.current_company["id"], new_ratio)
        self.note_text.delete("1.0", "end")
        self.note_text.insert("1.0", note)

    def save_ratio_note(self):
        if not self.current_company or not self.current_ratio:
            messagebox.showinfo(APP_TITLE, "Επίλεξε πρώτα δείκτη.")
            return
        self._save_current_ratio_note_silent()
        messagebox.showinfo(APP_TITLE, "Η σημείωση αποθηκεύτηκε.")

    def _pdf_footer(self, canvas, doc):
        canvas.saveState()
        font_regular, _ = register_pdf_fonts()
        page_width, page_height = canvas._pagesize
        canvas.setFont(font_regular, 6)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(page_width / 2, 0.65 * cm, "CORP Valuation app by G.Tsakalos")
        canvas.restoreState()

    def _make_pdf_chart(self, title, years, values, width=17.5 * cm, height=6.2 * cm, value_format="number"):
        drawing = Drawing(width, height)
        font_regular, font_bold = register_pdf_fonts()
        drawing.add(String(10, height - 14, title, fontName=font_bold, fontSize=12))
        chart = HorizontalLineChart()
        chart.x = 60
        chart.y = 35
        chart.height = height - 70
        chart.width = width - 95
        safe_values = [0 if v is None else round(float(v), 2) for v in values]
        chart.data = [safe_values]
        chart.categoryAxis.categoryNames = [str(y) for y in years]
        chart.categoryAxis.labels.boxAnchor = 'n'
        chart.categoryAxis.labels.angle = 0
        chart.joinedLines = 1
        chart.lines[0].strokeColor = colors.HexColor("#2F5597")
        chart.lines[0].strokeWidth = 2
        chart.lines[0].symbol = None
        min_val = min(safe_values) if safe_values else 0
        max_val = max(safe_values) if safe_values else 1
        if min_val == max_val:
            min_val -= 1
            max_val += 1
        pad = (max_val - min_val) * 0.15 or 1
        chart.valueAxis.valueMin = min_val - pad
        chart.valueAxis.valueMax = max_val + pad
        chart.valueAxis.valueStep = max((chart.valueAxis.valueMax - chart.valueAxis.valueMin) / 5.0, 1)
        if value_format == "percent":
            chart.valueAxis.labelTextFormat = lambda v: f"{v:.2f}%"
        elif value_format == "integer_thousands":
            chart.valueAxis.labelTextFormat = lambda v: format_thousands_dot(round(v))
        else:
            chart.valueAxis.labelTextFormat = "%.2f"
        drawing.add(chart)
        return drawing

    def _build_pdf(self, output_path: str, multi_company: bool = False):
        self._save_current_ratio_note_silent()
        if not REPORTLAB_OK:
            raise RuntimeError("Το reportlab δεν είναι διαθέσιμο.")
        font_regular, font_bold = register_pdf_fonts()
        styles = getSampleStyleSheet()
        styles["Title"].fontName = font_bold
        styles["Title"].fontSize = 16
        styles["Heading3"].fontName = font_bold
        styles["Heading3"].fontSize = 11
        styles["Normal"].fontName = font_regular
        styles["Normal"].fontSize = 9
        styles.add(ParagraphStyle(name="GreekSmall", fontName=font_regular, fontSize=7.2, leading=8.6))
        styles.add(ParagraphStyle(name="GreekSmallBold", fontName=font_bold, fontSize=7.4, leading=8.8))

        landscape_size = landscape(A4)
        portrait_size = A4
        doc = BaseDocTemplate(output_path, pagesize=landscape_size,
                              rightMargin=0.8 * cm, leftMargin=0.8 * cm,
                              topMargin=1.0 * cm, bottomMargin=1.1 * cm)

        landscape_frame = Frame(doc.leftMargin, doc.bottomMargin,
                                landscape_size[0] - doc.leftMargin - doc.rightMargin,
                                landscape_size[1] - doc.topMargin - doc.bottomMargin,
                                id="landscape_frame")
        portrait_frame = Frame(doc.leftMargin, doc.bottomMargin,
                               portrait_size[0] - doc.leftMargin - doc.rightMargin,
                               portrait_size[1] - doc.topMargin - doc.bottomMargin,
                               id="portrait_frame")

        doc.addPageTemplates([
            PageTemplate(id="landscape_report", pagesize=landscape_size, frames=[landscape_frame], onPage=self._pdf_footer),
            PageTemplate(id="portrait_charts", pagesize=portrait_size, frames=[portrait_frame], onPage=self._pdf_footer),
        ])

        story = []
        companies = self.app.db.all_companies_with_data() if multi_company else ([self.current_company] if self.current_company else [])
        if not companies:
            raise RuntimeError("Δεν υπάρχουν αναφορές για εξαγωγή.")

        first_company = True
        for company in companies:
            _, _, analyses = self.collect_company_analysis(company["id"])
            if not analyses:
                continue

            if not first_company:
                story.append(NextPageTemplate("landscape_report"))
                story.append(PageBreak())
            first_company = False

            story.append(Paragraph(company.get("name", ""), styles["Title"]))
            story.append(Spacer(1, 0.10 * cm))
            story.append(Paragraph(f"ΑΦΜ: {company.get('afm','')} | ΓΕΜΗ: {company.get('gemi','')}", styles["Normal"]))
            story.append(Spacer(1, 0.20 * cm))

            years = [a["year"] for a in analyses[:3]]
            header = ["Δείκτης"] + [str(y) for y in years] + ["Σημείωση"]
            data = [header]
            group_rows = []
            for item_type, label in iter_grouped_ratios():
                if item_type == "__GROUP__":
                    group_rows.append(len(data))
                    data.append([Paragraph(f"<b>{label}</b>", styles["GreekSmallBold"])] + [""] * len(years) + [""])
                    continue
                ratio_name = label
                row = [Paragraph(ratio_name, styles["GreekSmall"])]
                for analysis in analyses[:3]:
                    val = analysis["ratios"].get(ratio_name)
                    row.append(format_ratio_display(ratio_name, val))
                while len(row) < 1 + len(years):
                    row.append("")
                note = self.app.db.get_ratio_note(company["id"], ratio_name)
                row.append(Paragraph(note or "", styles["GreekSmall"]))
                data.append(row)

            usable_width = landscape_size[0] - doc.leftMargin - doc.rightMargin
            n_years = max(len(years), 1)
            first_col = 9.0 * cm
            note_col = 5.4 * cm
            year_col = (usable_width - first_col - note_col) / n_years
            col_widths = [first_col] + [year_col for _ in years] + [note_col]

            table = Table(data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e2f3")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("LEADING", (0, 0), (-1, -1), 8.6),
            ]
            for r in group_rows:
                style_cmds.extend([
                    ("BACKGROUND", (0, r), (-1, r), colors.HexColor("#e2f0d9")),
                    ("FONTNAME", (0, r), (-1, r), font_bold),
                    ("SPAN", (0, r), (-1, r)),
                ])
            table.setStyle(TableStyle(style_cmds))
            story.append(table)

            chart_years = [a["year"] for a in reversed(analyses[:3])]
            if chart_years:
                story.append(NextPageTemplate("portrait_charts"))
                story.append(PageBreak())
                story.append(Paragraph(f"{company.get('name', '')} - Charts", styles["Heading3"]))
                story.append(Spacer(1, 0.15 * cm))
                metric_map = [
                    ("EBIT", [a["ratios"].get("EBIT") for a in reversed(analyses[:3])], "integer_thousands"),
                    ("ROA", [None if a["ratios"].get("ROA (Απόδοση Συνόλου Ενεργητικού)") is None else a["ratios"].get("ROA (Απόδοση Συνόλου Ενεργητικού)") * 100 for a in reversed(analyses[:3])], "percent"),
                    ("ROE", [None if a["ratios"].get("ROE (Απόδοση Ι/Κ)") is None else a["ratios"].get("ROE (Απόδοση Ι/Κ)") * 100 for a in reversed(analyses[:3])], "percent"),
                ]
                for idx, (title, values, value_format) in enumerate(metric_map):
                    drawing = self._make_pdf_chart(title, chart_years, values, value_format=value_format)
                    story.append(drawing)
                    if idx < 2:
                        story.append(Spacer(1, 0.2 * cm))

        doc.build(story)

    def export_pdf(self):
        if not self.current_report_rows:
            self.run_analysis()
            if not self.current_report_rows:
                return
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")], title="Αποθήκευση PDF")
        if not path:
            return
        try:
            self._build_pdf(path, multi_company=False)
            messagebox.showinfo(APP_TITLE, f"Το PDF δημιουργήθηκε:\n{path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία export PDF:\n{e}")


    def export_xlsx(self):
        self._save_current_ratio_note_silent()
        if not self.current_report_rows:
            self.run_analysis()
            if not self.current_report_rows:
                return
        if not OPENPYXL_OK:
            messagebox.showerror(APP_TITLE, "Η βιβλιοθήκη openpyxl δεν είναι διαθέσιμη.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Αποθήκευση XLSX")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ratios"

            years = [a["year"] for a in self.current_report_rows[:3]]
            company_name = self.current_company.get("name", "") if self.current_company else ""
            ws.append(["Εταιρεία", company_name])
            ws.append(["ΑΦΜ", self.current_company.get("afm", "") if self.current_company else ""])
            ws.append(["ΓΕΜΗ", self.current_company.get("gemi", "") if self.current_company else ""])
            ws.append([])
            header_row = ws.max_row + 1
            ws.append(["Δείκτης"] + years + ["Σημείωση"])
            style_excel_header(ws, header_row)

            percent_ratios = PERCENT_RATIOS

            group_heading_rows = []
            for item_type, label in iter_grouped_ratios():
                if item_type == "__GROUP__":
                    ws.append([label] + [""] * len(years) + [""])
                    group_heading_rows.append(ws.max_row)
                    continue
                ratio_name = label
                row = [ratio_name]
                for analysis in self.current_report_rows[:3]:
                    row.append(analysis["ratios"].get(ratio_name))
                while len(row) < 1 + len(years):
                    row.append(None)
                row.append(self.app.db.get_ratio_note(self.current_company["id"], ratio_name) if self.current_company else "")
                ws.append(row)

            start_data_row = header_row + 1
            for r in range(start_data_row, ws.max_row + 1):
                ratio_name = ws.cell(r, 1).value
                for c in range(2, 2 + len(years)):
                    cell = ws.cell(r, c)
                    if ratio_name in percent_ratios and cell.value is not None:
                        cell.number_format = '0.00%'
                    elif ratio_name in DECIMAL4_RATIOS and cell.value is not None:
                        cell.number_format = '#,##0.0000'
                    else:
                        cell.number_format = '#,##0.00'
                ws.cell(r, 1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                ws.cell(r, 2 + len(years)).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

            for r in group_heading_rows:
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="E2F0D9")
            ws.freeze_panes = f"A{start_data_row}"
            ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
            autosize_worksheet(ws, min_width=12, max_width=42)

            totals_ws = wb.create_sheet("Totals")
            totals_ws.append(["Έτος", "EBIT", "EBITDA", "Depreciations (Αποσβέσεις Χρήσης)", "Working Capital Requirements (Ανάγκες σε Κ/Κ)"])
            style_excel_header(totals_ws, 1, "E2F0D9")
            for analysis in self.current_report_rows[:3]:
                totals_ws.append([
                    analysis["year"],
                    analysis["ratios"].get("EBIT"),
                    analysis["ratios"].get("EBITDA"),
                    analysis["ratios"].get("Depreciations (Αποσβέσεις Χρήσης)"),
                    analysis["ratios"].get("Working Capital Requirements (Ανάγκες σε Κ/Κ)"),
                ])
            for row in totals_ws.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    cell.number_format = '#,##0.00'
            autosize_worksheet(totals_ws)

            add_report_charts_sheet(wb, self.current_report_rows[:3])

            export_years = [a["year"] for a in self.current_report_rows[:3]]
            if self.current_company:
                append_statement_sheet_to_workbook(wb, "Ισολογισμός", self.current_company, export_years, self.app.db, "balance")
                append_statement_sheet_to_workbook(wb, "Αποτελέσματα Χρήσης", self.current_company, export_years, self.app.db, "income")

            wb.save(path)
            messagebox.showinfo(APP_TITLE, f"Το αρχείο XLSX δημιουργήθηκε:\n{path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία export XLSX:\n{e}")

    def export_multi_company_pdf(self):
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")], title="Αποθήκευση multi-report PDF")
        if not path:
            return
        try:
            self._build_pdf(path, multi_company=True)
            messagebox.showinfo(APP_TITLE, f"Το multi-report PDF δημιουργήθηκε:\n{path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία multi-report PDF:\n{e}")

    def print_report(self):
        if not self.current_report_rows:
            self.run_analysis()
            if not self.current_report_rows:
                return
        try:
            temp_pdf = os.path.join(tempfile.gettempdir(), "enterprise_value_report.pdf")
            self._build_pdf(temp_pdf, multi_company=False)
            if os.name == "nt":
                os.startfile(temp_pdf, "print")
            else:
                open_with_default_app(temp_pdf)
            messagebox.showinfo(APP_TITLE, "Η εντολή εκτύπωσης στάλθηκε.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία εκτύπωσης:\n{e}")



class CompareTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.sort_desc = {}
        self.current_rows = []
        self.build_ui()
        self.refresh_year_options()

    def build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text="Έτος:").pack(side="left")
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(top, textvariable=self.year_var, state="readonly", width=12)
        self.year_combo.pack(side="left", padx=5)
        ttk.Button(top, text="Υπολογισμός", command=self.run_compare).pack(side="left", padx=(10, 5))
        ttk.Button(top, text="Export", command=self.export_xlsx).pack(side="left", padx=5)

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree = ttk.Treeview(frame, show="headings")
        ysb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def clear_results(self):
        self.current_rows = []
        self.tree.delete(*self.tree.get_children())

    def refresh_year_options(self):
        years = self.app.db.all_years()
        self.year_combo["values"] = [str(y) for y in years]
        if years and not self.year_var.get():
            self.year_var.set(str(years[0]))

    def _format_compare_value(self, ratio_name, val):
        if val is None:
            return "-"
        return format_ratio_display(ratio_name, val)

    def run_compare(self):
        year_text = self.year_var.get().strip()
        if not year_text.isdigit():
            messagebox.showerror(APP_TITLE, "Επίλεξε έτος.")
            return
        year = int(year_text)
        rows = []
        for company in self.app.db.search_companies(""):
            bs = self.app.db.get_sheet("balance_sheets", company["id"], year)
            inc = self.app.db.get_sheet("income_sheets", company["id"], year)
            if not bs or not inc:
                continue
            prev_bs = self.app.db.get_sheet("balance_sheets", company["id"], year - 1)
            analysis = build_analysis(company, year, bs, inc, prev_bs if prev_bs else None)
            row = {"company": company.get("name", "")}
            row.update(analysis["ratios"])
            rows.append(row)
        self.current_rows = rows
        self.populate_tree(rows)

    def populate_tree(self, rows):
        cols = ["company"] + RATIO_ORDER
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        self.tree.heading("company", text="Επωνυμία", command=lambda c="company": self.sort_by(c))
        self.tree.column("company", width=220, anchor="w", stretch=False)
        for ratio_name in RATIO_ORDER:
            self.tree.heading(ratio_name, text=ratio_name, command=lambda c=ratio_name: self.sort_by(c))
            self.tree.column(ratio_name, width=150, anchor="center", stretch=False)
        for i, row in enumerate(rows):
            values = [row.get("company", "")]
            for ratio_name in RATIO_ORDER:
                values.append(self._format_compare_value(ratio_name, row.get(ratio_name)))
            self.tree.insert("", "end", iid=str(i), values=values)

    def sort_by(self, col):
        if not self.current_rows:
            return
        desc = self.sort_desc.get(col, False)
        if col == "company":
            self.current_rows.sort(key=lambda r: (r.get(col) or "").lower(), reverse=desc)
        else:
            self.current_rows.sort(key=lambda r: float('-inf') if r.get(col) is None else r.get(col), reverse=desc)
        self.sort_desc[col] = not desc
        self.populate_tree(self.current_rows)

    def export_xlsx(self):
        if not self.current_rows:
            self.run_compare()
            if not self.current_rows:
                return
        if not OPENPYXL_OK:
            messagebox.showerror(APP_TITLE, "Η βιβλιοθήκη openpyxl δεν είναι διαθέσιμη.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Αποθήκευση σύγκρισης XLSX")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Compare"
            year_text = self.year_var.get().strip()
            ws.append(["Σύγκριση εταιρειών", f"Έτος {year_text}"])
            ws.append([])
            header_row = ws.max_row + 1
            ws.append(["Επωνυμία"] + RATIO_ORDER)
            style_excel_header(ws, header_row)

            percent_ratios = PERCENT_RATIOS

            for row in self.current_rows:
                ws.append([row.get("company", "")] + [row.get(r) for r in RATIO_ORDER])

            for r in range(header_row + 1, ws.max_row + 1):
                for idx, ratio_name in enumerate(RATIO_ORDER, start=2):
                    cell = ws.cell(r, idx)
                    if ratio_name in percent_ratios and cell.value is not None:
                        cell.number_format = '0.00%'
                    elif ratio_name in DECIMAL4_RATIOS and cell.value is not None:
                        cell.number_format = '#,##0.0000'
                    else:
                        cell.number_format = '#,##0.00'

            ws.freeze_panes = f"A{header_row + 1}"
            ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
            autosize_worksheet(ws, min_width=12, max_width=28)

            meta = wb.create_sheet("Meta")
            meta.append(["Έτος", year_text])
            meta.append(["Αριθμός εταιρειών", len(self.current_rows)])
            autosize_worksheet(meta)

            wb.save(path)
            messagebox.showinfo(APP_TITLE, f"Το αρχείο XLSX δημιουργήθηκε:\n{path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία export XLSX:\n{e}")

class App(tk.Tk):
    def __init__(self, db_path: str, cfg_path: str):
        super().__init__()
        self.cfg_path = cfg_path
        self.title(APP_TITLE)
        self.geometry("1460x900")
        self.minsize(1240, 760)
        self.db = SQLiteDB(db_path)
        self.style = ttk.Style(self)
        if "vista" in self.style.theme_names():
            self.style.theme_use("vista")

        self.bind_class("Treeview", "<MouseWheel>", self._on_treeview_mousewheel, add="+")
        self.bind_class("Treeview", "<Button-4>", self._on_treeview_mousewheel, add="+")
        self.bind_class("Treeview", "<Button-5>", self._on_treeview_mousewheel, add="+")
        self.bind_class("Text", "<MouseWheel>", self._on_text_mousewheel, add="+")
        self.bind_class("Text", "<Button-4>", self._on_text_mousewheel, add="+")
        self.bind_class("Text", "<Button-5>", self._on_text_mousewheel, add="+")

        menubar = tk.Menu(self)
        app_menu = tk.Menu(menubar, tearoff=0)
        app_menu.add_command(label="OpenDB", command=self.open_db)
        app_menu.add_command(label="SaveDB", command=self.save_db)
        app_menu.add_separator()
        app_menu.add_command(label="ConfigFileEdit", command=self.edit_config_file)
        app_menu.add_command(label="IndexMap", command=self.open_index_map)
        menubar.add_cascade(label="Menu", menu=app_menu)
        self.config(menu=menubar)

        info = ttk.Frame(self)
        info.pack(fill="x", padx=10, pady=(10, 0))
        self.db_label = ttk.Label(info, text=f"DB: {Path(db_path).resolve()}")
        self.db_label.pack(side="left")

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.company_tab = CompanyTab(notebook, self)
        self.balance_tab = BalanceSheetTab(notebook, self)
        self.income_tab = IncomeSheetTab(notebook, self)
        self.analysis_tab = AnalysisTab(notebook, self)
        self.compare_tab = CompareTab(notebook, self)

        notebook.add(self.company_tab, text="1. Companies")
        notebook.add(self.balance_tab, text="2. Ισολογισμός")
        notebook.add(self.income_tab, text="3. Αποτελέσματα Χρήσης")
        notebook.add(self.analysis_tab, text="4. Δείκτες / Report")
        notebook.add(self.compare_tab, text="5. Compare")

        self.refresh_company_combos()
        self.protocol("WM_DELETE_WINDOW", self.on_close)


    def _on_treeview_mousewheel(self, event):
        widget = event.widget
        if event.delta:
            widget.yview_scroll(int(-event.delta / 120), "units")
        elif getattr(event, "num", None) == 4:
            widget.yview_scroll(-3, "units")
        elif getattr(event, "num", None) == 5:
            widget.yview_scroll(3, "units")
        return "break"

    def _on_text_mousewheel(self, event):
        widget = event.widget
        if event.delta:
            widget.yview_scroll(int(-event.delta / 120), "units")
        elif getattr(event, "num", None) == 4:
            widget.yview_scroll(-3, "units")
        elif getattr(event, "num", None) == 5:
            widget.yview_scroll(3, "units")
        return "break"

    def on_close(self):
        try:
            self.db.close()
        finally:
            self.destroy()

    def company_display(self, comp: dict) -> str:
        return f"{comp.get('name','')} | ΑΦΜ: {comp.get('afm','')} | ΓΕΜΗ: {comp.get('gemi','')}"

    def refresh_company_combos(self):
        options = [(c["id"], self.company_display(c)) for c in self.db.search_companies("")]
        if hasattr(self, "balance_tab"):
            self.balance_tab.set_company_options(options)
        if hasattr(self, "income_tab"):
            self.income_tab.set_company_options(options)
        if hasattr(self, "analysis_tab"):
            self.analysis_tab.set_company_options(options)
        if hasattr(self, "compare_tab"):
            self.compare_tab.refresh_year_options()

    def ask_company_for_import(self) -> int:
        options = self.db.search_companies("")
        if options:
            names = "\n".join([f"- {c['name']}" for c in options[:20]])
            text = simpledialog.askstring(
                APP_TITLE,
                "Δώσε ακριβώς την επωνυμία εταιρείας για import.\nΑν δεν υπάρχει, γράψε νέα επωνυμία για δημιουργία νέας εγγραφής.\n\nΔιαθέσιμες εταιρείες:\n" + names,
                parent=self,
            )
        else:
            text = simpledialog.askstring(APP_TITLE, "Δώσε την επωνυμία της εταιρείας για import:", parent=self)
        if not text:
            raise ImportCanceled()
        text = text.strip()
        for company in options:
            if company["name"].strip().lower() == text.lower():
                return company["id"]
        afm = simpledialog.askstring(APP_TITLE, "ΑΦΜ νέας εταιρείας (προαιρετικό):", parent=self) or ""
        gemi = simpledialog.askstring(APP_TITLE, "ΓΕΜΗ νέας εταιρείας (προαιρετικό):", parent=self) or ""
        payload = {"name": text, "afm": afm.strip(), "gemi": gemi.strip(), "gemi_url": "", "company_url": "", "notes": ""}
        saved = self.db.upsert_company(payload)
        self.refresh_company_combos()
        self.company_tab.refresh_list()
        return saved["id"]

    def reopen_db(self, path: str):
        try:
            self.db.close()
        except Exception:
            pass
        self.db = SQLiteDB(path)
        self.company_tab.app = self.balance_tab.app = self.income_tab.app = self.analysis_tab.app = self.compare_tab.app = self
        self.company_tab.refresh_list()
        self.balance_tab.refresh_records()
        self.income_tab.refresh_records()
        self.refresh_company_combos()
        if hasattr(self, "compare_tab"):
            self.compare_tab.refresh_year_options()
        self.db_label.configure(text=f"DB: {Path(path).resolve()}")

    def open_db(self):
        path = filedialog.askopenfilename(filetypes=[("SQLite", "*.sqlite *.db")], title="Άνοιγμα βάσης SQLite")
        if not path:
            return
        self.reopen_db(path)

    def save_db(self):
        try:
            self.db.conn.commit()
            messagebox.showinfo(APP_TITLE, "Η βάση αποθηκεύτηκε.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αποτυχία αποθήκευσης:\n{e}")

    def edit_config_file(self):
        try:
            ensure_default_cfg(self.cfg_path)
            open_with_default_app(self.cfg_path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αδυναμία ανοίγματος config file:\n{e}")

    def open_index_map(self):
        try:
            path = get_cfg_value(self.cfg_path, "app", "index_map_path", resource_path(DEFAULT_INDEX_MAP))
            open_with_default_app(path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Αδυναμία ανοίγματος IndexMap:\n{e}")





def main():
    cfg_path = resource_path(DEFAULT_CFG)
    ensure_default_cfg(cfg_path)
    db_path = get_cfg_value(cfg_path, "app", "db_path", resource_path(DEFAULT_DB))
    app = App(db_path, cfg_path)
    app.mainloop()


if __name__ == "__main__":
    main()
